"""
AdSellix Audit Tool - Core Data Loaders
Version: 1.0.0
Status: PROTECTED - Do not modify without backup

This module handles parsing of all Amazon report formats.
"""

import pandas as pd
import io
import re
from typing import Dict, Tuple, Optional, List
from datetime import datetime


# =============================================================================
# MARKETPLACE CONFIGURATION
# =============================================================================

MARKETPLACES = {
    "US": {"name": "United States", "currency": "USD", "domain": "amazon.com"},
    "CA": {"name": "Canada", "currency": "CAD", "domain": "amazon.ca"},
    "UK": {"name": "United Kingdom", "currency": "GBP", "domain": "amazon.co.uk"},
    "DE": {"name": "Germany", "currency": "EUR", "domain": "amazon.de"},
    "FR": {"name": "France", "currency": "EUR", "domain": "amazon.fr"},
    "IT": {"name": "Italy", "currency": "EUR", "domain": "amazon.it"},
    "ES": {"name": "Spain", "currency": "EUR", "domain": "amazon.es"},
    "MX": {"name": "Mexico", "currency": "MXN", "domain": "amazon.com.mx"},
    "AU": {"name": "Australia", "currency": "AUD", "domain": "amazon.com.au"},
    "JP": {"name": "Japan", "currency": "JPY", "domain": "amazon.co.jp"},
}


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def clean_numeric(value) -> float:
    """Convert various numeric formats to float."""
    if pd.isna(value) or value == '' or value == '-' or value == '--':
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    # Remove currency symbols, commas, percentage signs
    cleaned = str(value).replace('$', '').replace('£', '').replace('€', '')
    cleaned = cleaned.replace(',', '').replace('%', '').strip()
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def clean_percentage(value) -> float:
    """Convert percentage strings to decimal (e.g., '15.5%' -> 0.155)."""
    num = clean_numeric(value)
    # If the value was already a percentage string, it's now just the number
    # We keep it as-is since reports store percentages as numbers (15.5 = 15.5%)
    return num


def parse_metadata_row(row_text: str) -> Dict:
    """Parse the metadata row from Brand Analytics reports."""
    metadata = {}
    # Pattern: Key=["Value"]
    pattern = r'(\w+(?:\s+\w+)?)\s*=\s*\["([^"]+)"\]'
    matches = re.findall(pattern, row_text)
    for key, value in matches:
        key_clean = key.strip().lower().replace(' ', '_')
        metadata[key_clean] = value
    return metadata


def detect_file_encoding(file_content: bytes) -> str:
    """Detect file encoding, handling BOM."""
    if file_content.startswith(b'\xef\xbb\xbf'):
        return 'utf-8-sig'
    return 'utf-8'


# =============================================================================
# SQP BRAND VIEW LOADER
# =============================================================================

def load_sqp_brand_view(file) -> Tuple[pd.DataFrame, Dict]:
    """
    Load Search Query Performance - Brand View report.
    
    Returns:
        Tuple of (DataFrame with data, Dict with metadata)
    """
    # Read file content
    if hasattr(file, 'read'):
        content = file.read()
        if isinstance(content, bytes):
            content = content.decode('utf-8-sig')
        file.seek(0)
    else:
        with open(file, 'r', encoding='utf-8-sig') as f:
            content = f.read()
    
    lines = content.strip().split('\n')
    
    # Parse metadata from first row
    metadata = parse_metadata_row(lines[0])
    
    # Read data starting from row 2 (index 1)
    data_content = '\n'.join(lines[1:])
    df = pd.read_csv(io.StringIO(data_content))
    
    # Clean column names
    df.columns = df.columns.str.strip().str.replace('"', '')
    
    # Convert numeric columns
    numeric_cols = [
        'Search Query Score', 'Search Query Volume',
        'Impressions: Total Count', 'Impressions: Brand Count', 'Impressions: Brand Share %',
        'Clicks: Total Count', 'Clicks: Click Rate %', 'Clicks: Brand Count', 'Clicks: Brand Share %',
        'Cart Adds: Total Count', 'Cart Adds: Cart Add Rate %', 'Cart Adds: Brand Count', 'Cart Adds: Brand Share %',
        'Purchases: Total Count', 'Purchases: Purchase Rate %', 'Purchases: Brand Count', 'Purchases: Brand Share %'
    ]
    
    for col in numeric_cols:
        if col in df.columns:
            df[col] = df[col].apply(clean_numeric)
    
    # Price columns
    price_cols = [c for c in df.columns if 'Price' in c]
    for col in price_cols:
        if col in df.columns:
            df[col] = df[col].apply(clean_numeric)
    
    return df, metadata


# =============================================================================
# SQP ASIN VIEW LOADER
# =============================================================================

def load_sqp_asin_view(file) -> Tuple[pd.DataFrame, Dict]:
    """
    Load Search Query Performance - ASIN View report.
    
    Returns:
        Tuple of (DataFrame with data, Dict with metadata including ASIN)
    """
    # Read file content
    if hasattr(file, 'read'):
        content = file.read()
        if isinstance(content, bytes):
            content = content.decode('utf-8-sig')
        file.seek(0)
    else:
        with open(file, 'r', encoding='utf-8-sig') as f:
            content = f.read()
    
    lines = content.strip().split('\n')
    
    # Parse metadata from first row (includes ASIN)
    metadata = parse_metadata_row(lines[0])
    
    # Extract ASIN from metadata
    if 'asin_or_product' in metadata:
        metadata['asin'] = metadata['asin_or_product']
    
    # Read data starting from row 2
    data_content = '\n'.join(lines[1:])
    df = pd.read_csv(io.StringIO(data_content))
    
    # Clean column names
    df.columns = df.columns.str.strip().str.replace('"', '')
    
    # Convert numeric columns (same as brand view but with ASIN instead of Brand)
    numeric_cols = [
        'Search Query Score', 'Search Query Volume',
        'Impressions: Total Count', 'Impressions: ASIN Count', 'Impressions: ASIN Share %',
        'Clicks: Total Count', 'Clicks: Click Rate %', 'Clicks: ASIN Count', 'Clicks: ASIN Share %',
        'Cart Adds: Total Count', 'Cart Adds: Cart Add Rate %', 'Cart Adds: ASIN Count', 'Cart Adds: ASIN Share %',
        'Purchases: Total Count', 'Purchases: Purchase Rate %', 'Purchases: ASIN Count', 'Purchases: ASIN Share %'
    ]
    
    for col in numeric_cols:
        if col in df.columns:
            df[col] = df[col].apply(clean_numeric)
    
    return df, metadata


# =============================================================================
# BUSINESS REPORT LOADER
# =============================================================================

def load_business_report(file) -> pd.DataFrame:
    """
    Load Business Report - Detail Page Sales and Traffic.
    
    Returns:
        DataFrame with cleaned data
    """
    df = pd.read_csv(file, encoding='utf-8-sig')
    
    # Clean column names
    df.columns = df.columns.str.strip()
    
    # Numeric columns to clean
    numeric_cols = [
        'Sessions - Total', 'Page Views - Total', 'Page Views - Total - B2B',
        'Units Ordered', 'Units Ordered - B2B',
        'Total Order Items', 'Total Order Items - B2B'
    ]
    
    for col in numeric_cols:
        if col in df.columns:
            df[col] = df[col].apply(clean_numeric)
    
    # Percentage columns
    pct_cols = [
        'Session Percentage - Total', 'Session Percentage - Total - B2B',
        'Page Views Percentage - Total', 'Page Views Percentage - Total - B2B',
        'Featured Offer (Buy Box) Percentage', 'Featured Offer (Buy Box) Percentage - B2B',
        'Unit Session Percentage', 'Unit Session Percentage - B2B'
    ]
    
    for col in pct_cols:
        if col in df.columns:
            df[col] = df[col].apply(clean_percentage)
    
    # Sales columns (currency)
    sales_cols = ['Ordered Product Sales', 'Ordered Product Sales - B2B']
    for col in sales_cols:
        if col in df.columns:
            df[col] = df[col].apply(clean_numeric)
    
    return df


# =============================================================================
# PPC BULK SHEET LOADER
# =============================================================================

def load_ppc_bulk_sheet(file) -> Dict[str, pd.DataFrame]:
    """
    Load PPC Bulk Sheet (Excel format with multiple sheets).
    
    Returns:
        Dict with sheet names as keys and DataFrames as values
    """
    import openpyxl
    
    sheets_to_load = [
        'Sponsored Products Campaigns',
        'Sponsored Brands Campaigns', 
        'Sponsored Display Campaigns',
        'SP Search Term Report',
        'SB Search Term Report'
    ]
    
    result = {}
    
    # Load workbook
    if hasattr(file, 'read'):
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    else:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    
    for sheet_name in sheets_to_load:
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            data = list(sheet.iter_rows(values_only=True))
            
            if len(data) > 1:
                headers = data[0]
                df = pd.DataFrame(data[1:], columns=headers)
                
                # Convert numeric columns
                numeric_cols = [
                    'Impressions', 'Clicks', 'Spend', 'Sales', 'Orders', 'Units',
                    'Bid', 'Daily Budget', 'Ad Group Default Bid', 'Percentage'
                ]
                
                for col in numeric_cols:
                    if col in df.columns:
                        df[col] = df[col].apply(clean_numeric)
                
                # Percentage columns
                pct_cols = ['Click-through Rate', 'Conversion Rate', 'ACOS', 'CPC', 'ROAS']
                for col in pct_cols:
                    if col in df.columns:
                        df[col] = df[col].apply(clean_numeric)
                
                result[sheet_name] = df
    
    wb.close()
    return result


# =============================================================================
# INVENTORY REPORT LOADER
# =============================================================================

def load_inventory_report(file) -> pd.DataFrame:
    """
    Load FBA Inventory Report.
    
    Returns:
        DataFrame with cleaned inventory data
    """
    df = pd.read_csv(file, encoding='utf-8-sig')
    
    # Clean column names
    df.columns = df.columns.str.strip()
    
    # Key numeric columns
    numeric_cols = [
        'available', 'pending-removal-quantity',
        'inv-age-0-to-90-days', 'inv-age-91-to-180-days', 
        'inv-age-181-to-270-days', 'inv-age-271-to-365-days',
        'inv-age-366-to-455-days', 'inv-age-456-plus-days',
        'units-shipped-t7', 'units-shipped-t30', 'units-shipped-t60', 'units-shipped-t90',
        'days-of-supply', 'estimated-excess-quantity',
        'weeks-of-cover-t30', 'weeks-of-cover-t90',
        'sell-through', 'estimated-storage-cost-next-month',
        'inbound-quantity', 'inbound-working', 'inbound-shipped', 'inbound-received'
    ]
    
    for col in numeric_cols:
        if col in df.columns:
            df[col] = df[col].apply(clean_numeric)
    
    # Price columns
    price_cols = ['your-price', 'sales-price', 'lowest-price-new-plus-shipping', 
                  'lowest-price-used', 'featuredoffer-price']
    for col in price_cols:
        if col in df.columns:
            df[col] = df[col].apply(clean_numeric)
    
    return df


# =============================================================================
# COGS LOADER
# =============================================================================

def load_cogs(file) -> pd.DataFrame:
    """
    Load COGS (Cost of Goods Sold) data.
    
    Returns:
        DataFrame with ASIN/SKU level costs
    """
    df = pd.read_csv(file, encoding='utf-8-sig')
    
    # Clean column names
    df.columns = df.columns.str.strip()
    
    # Convert cost columns
    cost_cols = ['Cost', 'ShippingCostPerOrder', 'SurchargeForShippingAbroad', 
                 'Value_of_unsellable_returns']
    
    for col in cost_cols:
        if col in df.columns:
            df[col] = df[col].apply(clean_numeric)
    
    return df


# =============================================================================
# FBA FEES LOADER
# =============================================================================

def load_fba_fees(file) -> pd.DataFrame:
    """
    Load FBA and Referral Fees report.
    
    Returns:
        DataFrame with fee data per ASIN/SKU
    """
    df = pd.read_csv(file, encoding='utf-8-sig')
    
    # Clean column names
    df.columns = df.columns.str.strip()
    
    # Fee columns
    fee_cols = [
        'your-price', 'sales-price',
        'estimated-fee-total', 'estimated-referral-fee-per-unit',
        'estimated-variable-closing-fee', 'expected-domestic-fulfilment-fee-per-unit'
    ]
    
    # EFN fee columns
    efn_cols = [c for c in df.columns if 'expected-efn-fulfilment-fee' in c]
    fee_cols.extend(efn_cols)
    
    for col in fee_cols:
        if col in df.columns:
            df[col] = df[col].apply(clean_numeric)
    
    return df


# =============================================================================
# CATEGORY LISTING REPORT LOADER
# =============================================================================

def load_category_listing_report(file) -> pd.DataFrame:
    """
    Load Category Listing Report (XLSM format).
    Extracts parent-child hierarchy.
    
    Returns:
        DataFrame with hierarchy data
    """
    import openpyxl
    
    if hasattr(file, 'read'):
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    else:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    
    sheet = wb['Template']
    data = list(sheet.iter_rows(values_only=True))
    
    # Row 4 is headers (index 3)
    headers = data[3]
    
    # Data starts at row 7 (index 6)
    rows = data[6:]
    
    # Key columns we need (based on analysis)
    col_indices = {
        'Status': 0,
        'SKU': 2,
        'Item Name': 5,
        'Brand Name': 6,
        'Product Id': 8,
        'Colour': 100,
        'Size': 101,
        'Parentage Level': 287,
        'Parent SKU': 288,
        'Variation Theme': 289
    }
    
    # Extract only relevant columns
    extracted_data = []
    for row in rows:
        if row and len(row) > 289:
            row_data = {}
            for col_name, idx in col_indices.items():
                row_data[col_name] = row[idx] if idx < len(row) else None
            # Only include rows with SKU
            if row_data['SKU']:
                extracted_data.append(row_data)
    
    wb.close()
    
    df = pd.DataFrame(extracted_data)
    return df


# =============================================================================
# TOP SEARCH TERMS LOADER
# =============================================================================

def load_top_search_terms(file) -> Tuple[pd.DataFrame, Dict]:
    """
    Load Top Search Terms report (Market Intelligence).
    
    Returns:
        Tuple of (DataFrame with data, Dict with metadata)
    """
    if hasattr(file, 'read'):
        content = file.read()
        if isinstance(content, bytes):
            content = content.decode('utf-8-sig')
        file.seek(0)
    else:
        with open(file, 'r', encoding='utf-8-sig') as f:
            content = f.read()
    
    lines = content.strip().split('\n')
    
    # Parse metadata from first row
    metadata = parse_metadata_row(lines[0])
    
    # Read data starting from row 2
    data_content = '\n'.join(lines[1:])
    df = pd.read_csv(io.StringIO(data_content))
    
    # Clean column names
    df.columns = df.columns.str.strip().str.replace('"', '')
    
    # Convert numeric columns
    df['Search Frequency Rank'] = df['Search Frequency Rank'].apply(clean_numeric)
    
    # Click and conversion share columns
    share_cols = [c for c in df.columns if 'Share' in c]
    for col in share_cols:
        df[col] = df[col].apply(clean_numeric)
    
    return df, metadata


# =============================================================================
# SEARCH CATALOG PERFORMANCE LOADER
# =============================================================================

def load_search_catalog_performance(file) -> Tuple[pd.DataFrame, Dict]:
    """
    Load Search Catalog Performance report.
    
    Returns:
        Tuple of (DataFrame with funnel data, Dict with metadata)
    """
    if hasattr(file, 'read'):
        content = file.read()
        if isinstance(content, bytes):
            content = content.decode('utf-8-sig')
        file.seek(0)
    else:
        with open(file, 'r', encoding='utf-8-sig') as f:
            content = f.read()
    
    lines = content.strip().split('\n')
    
    # Parse metadata
    metadata = parse_metadata_row(lines[0])
    
    # Read data
    data_content = '\n'.join(lines[1:])
    df = pd.read_csv(io.StringIO(data_content))
    
    # Clean column names
    df.columns = df.columns.str.strip().str.replace('"', '')
    
    # Convert all numeric columns
    numeric_cols = [c for c in df.columns if any(x in c for x in 
                   ['Impressions', 'Clicks', 'Basket', 'Purchases', 'Sales', 'Rate', 'Speed'])]
    
    for col in numeric_cols:
        if col in df.columns:
            df[col] = df[col].apply(clean_numeric)
    
    return df, metadata


# =============================================================================
# RETURNS REPORT LOADER
# =============================================================================

def load_returns_report(file) -> pd.DataFrame:
    """
    Load Returns Report (TSV format).
    
    Returns:
        DataFrame with returns data
    """
    df = pd.read_csv(file, sep='\t', encoding='utf-8-sig')
    
    # Clean column names
    df.columns = df.columns.str.strip()
    
    # Convert date columns
    date_cols = ['Order date', 'Return request date', 'Return delivery date']
    for col in date_cols:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors='coerce')
    
    # Numeric columns
    numeric_cols = ['Return quantity', 'Order quantity', 'Refunded amount', 'Order amount']
    for col in numeric_cols:
        if col in df.columns:
            df[col] = df[col].apply(clean_numeric)
    
    return df


# =============================================================================
# MASTER DATA AGGREGATOR
# =============================================================================

class AuditDataLoader:
    """
    Master class for loading all audit data sources.
    """
    
    def __init__(self, marketplace: str = "US"):
        self.marketplace = marketplace
        self.marketplace_info = MARKETPLACES.get(marketplace, MARKETPLACES["US"])
        self.data = {}
        self.metadata = {}
    
    def load_sqp_brand_current(self, file):
        """Load current period SQP Brand View."""
        df, meta = load_sqp_brand_view(file)
        self.data['sqp_brand_current'] = df
        self.metadata['sqp_brand_current'] = meta
        return df
    
    def load_sqp_brand_previous(self, file):
        """Load previous period SQP Brand View."""
        df, meta = load_sqp_brand_view(file)
        self.data['sqp_brand_previous'] = df
        self.metadata['sqp_brand_previous'] = meta
        return df
    
    def load_sqp_asin(self, file, asin: str = None):
        """Load ASIN-level SQP data."""
        df, meta = load_sqp_asin_view(file)
        asin_key = asin or meta.get('asin', 'unknown')
        
        if 'sqp_asin' not in self.data:
            self.data['sqp_asin'] = {}
            self.metadata['sqp_asin'] = {}
        
        self.data['sqp_asin'][asin_key] = df
        self.metadata['sqp_asin'][asin_key] = meta
        return df
    
    def load_business_report(self, file):
        """Load Business Report."""
        df = load_business_report(file)
        self.data['business_report'] = df
        return df
    
    def load_ppc(self, file):
        """Load PPC Bulk Sheet."""
        sheets = load_ppc_bulk_sheet(file)
        self.data['ppc'] = sheets
        return sheets
    
    def load_inventory(self, file):
        """Load Inventory Report."""
        df = load_inventory_report(file)
        self.data['inventory'] = df
        return df
    
    def load_cogs(self, file):
        """Load COGS data."""
        df = load_cogs(file)
        self.data['cogs'] = df
        return df
    
    def load_fba_fees(self, file):
        """Load FBA Fees."""
        df = load_fba_fees(file)
        self.data['fba_fees'] = df
        return df
    
    def load_clr(self, file):
        """Load Category Listing Report."""
        df = load_category_listing_report(file)
        self.data['clr'] = df
        return df
    
    def get_all_data(self) -> Dict:
        """Return all loaded data."""
        return self.data
    
    def get_all_metadata(self) -> Dict:
        """Return all metadata."""
        return self.metadata
